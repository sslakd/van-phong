#!/usr/bin/env python3
"""Generate all office template files with unified green theme style."""
import os, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = os.path.expanduser("~/Desktop/Openclaw/17_Office_Templates/templates/excel")
os.makedirs(OUT, exist_ok=True)

# ===== Style constants =====
GREEN_DARK = "14532d"
GREEN_MID = "16a34a"
GREEN_LIGHT = "dcfce7"
GREEN_PALE = "f0fdf4"
WHITE = "ffffff"

font_title = Font(name="Inter", size=14, bold=True, color=GREEN_DARK)
font_header = Font(name="Inter", size=11, bold=True, color=WHITE)
font_body = Font(name="Inter", size=10, color=GREEN_DARK)
font_body_bold = Font(name="Inter", size=10, bold=True, color=GREEN_DARK)

fill_header = PatternFill(start_color=GREEN_MID, end_color=GREEN_MID, fill_type="solid")
fill_light = PatternFill(start_color=GREEN_LIGHT, end_color=GREEN_LIGHT, fill_type="solid")
fill_pale = PatternFill(start_color=GREEN_PALE, end_color=GREEN_PALE, fill_type="solid")
fill_white = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")

align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
align_right = Alignment(horizontal="right", vertical="center")

thin_border = Border(
    left=Side(style="thin", color="bbf7d0"),
    right=Side(style="thin", color="bbf7d0"),
    top=Side(style="thin", color="bbf7d0"),
    bottom=Side(style="thin", color="bbf7d0"),
)
header_border = Border(
    left=Side(style="thin", color=GREEN_MID),
    right=Side(style="thin", color=GREEN_MID),
    top=Side(style="thin", color=GREEN_MID),
    bottom=Side(style="medium", color=GREEN_MID),
)

def style_header_row(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = header_border

def style_body_cell(cell, center=False):
    cell.font = font_body
    cell.fill = fill_white
    cell.alignment = align_center if center else align_left
    cell.border = thin_border

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# =====================================================
# 1. Personal Budget Tracker
# =====================================================
def create_budget():
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Ngân sách"
    ws.merge_cells("A1:F1")
    ws["A1"].value = "Bảng Ngân Sách Cá Nhân"
    ws["A1"].font = font_title; ws["A1"].alignment = align_center
    ws.row_dimensions[1].height = 36

    headers = ["Mục", "Ngân sách (VNĐ)", "Đã chi (VNĐ)", "Còn lại (VNĐ)", "% Đã dùng", "Ghi chú"]
    for c, h in enumerate(headers, 1): ws.cell(row=3, column=c, value=h)
    style_header_row(ws, 3, 6)

    items = [("Nhà ở (thuê/thế chấp)", 5000000, 5000000, "Thuê nhà T5"),
             ("Điện, nước, internet", 1500000, 1200000, ""),
             ("Ăn uống", 3000000, 2800000, ""),
             ("Di chuyển (xăng/giữ xe)", 1000000, 800000, ""),
             ("Bảo hiểm", 2000000, 2000000, ""),
             ("Giải trí", 1000000, 500000, ""),
             ("Tiết kiệm", 2000000, 1500000, ""),
             ("Khác", 500000, 300000, "")]
    for i, (name, budget, spent, note) in enumerate(items, 4):
        ws.cell(row=i, column=1, value=name)
        ws.cell(row=i, column=2, value=budget)
        ws.cell(row=i, column=3, value=spent)
        ws.cell(row=i, column=4, value=budget - spent)
        ws.cell(row=i, column=5, value=f"{round(spent/budget*100)}%")
        ws.cell(row=i, column=6, value=note)
        for c in range(1, 7):
            style_body_cell(ws.cell(row=i, column=c), center=(c > 1))
            if i % 2 == 0: ws.cell(row=i, column=c).fill = fill_pale

    row_t = 4 + len(items)
    ws.cell(row=row_t, column=1, value="TỔNG CỘNG").font = font_body_bold
    for col in [2,3,4]:
        cl = get_column_letter(col)
        ws.cell(row=row_t, column=col, value=f"=SUM({cl}4:{cl}{row_t-1})")
        ws.cell(row=row_t, column=col).font = font_body_bold
        ws.cell(row=row_t, column=col).alignment = align_center
    ws.cell(row=row_t, column=5, value="100%").font = font_body_bold
    ws.cell(row=row_t, column=5).alignment = align_center
    set_col_widths(ws, [30, 18, 18, 18, 14, 25])
    ws.sheet_view.showGridLines = False
    wb.save(os.path.join(OUT, "personal-budget.xlsx"))
    print("✅ personal-budget.xlsx")

# =====================================================
# 2. Invoice
# =====================================================
def create_invoice():
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Hóa đơn"
    ws.merge_cells("A1:F1")
    ws["A1"].value = "HÓA ĐƠN BÁN HÀNG"
    ws["A1"].font = Font(name="Inter", size=18, bold=True, color=GREEN_DARK)
    ws["A1"].alignment = align_center; ws.row_dimensions[1].height = 44
    ws["A3"].value = "Số hóa đơn: INV-2025-001"; ws["A3"].font = font_body_bold
    ws["E3"].value = "Ngày: 21/05/2025"; ws["E3"].font = font_body_bold; ws["E3"].alignment = align_right
    ws["A4"].value = "Khách hàng: Công ty TNHH ABC"; ws["A4"].font = font_body
    ws["A5"].value = "Địa chỉ: 123 Nguyễn Huệ, Q.1, TP.HCM"; ws["A5"].font = font_body
    headers = ["STT", "Sản phẩm / Dịch vụ", "Đơn giá (VNĐ)", "Số lượng", "Thành tiền (VNĐ)", "Thuế VAT"]
    for c, h in enumerate(headers, 1): ws.cell(row=7, column=c, value=h)
    style_header_row(ws, 7, 6)
    items = [(1,"Thiết kế website",15000000,1),(2,"Viết content SEO",3000000,5),(3,"Chạy quảng cáo Google Ads",5000000,1),(4,"Hosting + Domain 12 tháng",2400000,1)]
    for i,(stt,name,price,qty) in enumerate(items,8):
        total=price*qty; vat=int(total*0.1)
        ws.cell(row=i, column=1, value=stt); ws.cell(row=i,column=2,value=name)
        ws.cell(row=i,column=3,value=price); ws.cell(row=i,column=4,value=qty)
        ws.cell(row=i,column=5,value=total); ws.cell(row=i,column=6,value=vat)
        for c in range(1,7):
            style_body_cell(ws.cell(row=i,column=c),center=(c!=2))
            if i%2==0: ws.cell(row=i,column=c).fill=fill_pale
    row_t=8+len(items)
    ws.cell(row=row_t,column=2,value="TỔNG CỘNG").font=font_body_bold
    for col in[5,6]:
        cl=get_column_letter(col); ws.cell(row=row_t,column=col,value=f"=SUM({cl}8:{cl}{row_t-1})")
        ws.cell(row=row_t,column=col).font=font_body_bold; ws.cell(row=row_t,column=col).alignment=align_center
    ws.merge_cells(f"A{row_t+1}:D{row_t+1}")
    ws.cell(row=row_t+1,column=1,value="Tổng tiền (đã bao gồm VAT 10%):").font=font_body_bold; ws.cell(row=row_t+1,column=1).alignment=align_right
    ws.cell(row=row_t+1,column=5,value=f"={get_column_letter(5)}{row_t}+{get_column_letter(6)}{row_t}")
    ws.cell(row=row_t+1,column=5).font=Font(name="Inter",size=12,bold=True,color=GREEN_MID); ws.cell(row=row_t+1,column=5).alignment=align_center
    ws.merge_cells(f"A{row_t+3}:F{row_t+3}")
    ws.cell(row=row_t+3,column=1,value="Cảm ơn quý khách!").font=Font(name="Inter",size=10,italic=True,color="166534"); ws.cell(row=row_t+3,column=1).alignment=align_center
    set_col_widths(ws,[6,35,18,10,18,14]); ws.sheet_view.showGridLines=False
    wb.save(os.path.join(OUT, "invoice.xlsx")); print("✅ invoice.xlsx")

# =====================================================
# 3. Timesheet
# =====================================================
def create_timesheet():
    wb=openpyxl.Workbook(); ws=wb.active; ws.title="Chấm công"
    ws.merge_cells("A1:G1"); ws["A1"].value="BẢNG CHẤM CÔNG THÁNG 5/2025"
    ws["A1"].font=font_title; ws["A1"].alignment=align_center; ws.row_dimensions[1].height=36
    ws["A3"].value="Nhân viên: Nguyễn Văn A"; ws["A3"].font=font_body_bold
    ws["D3"].value="Phòng: Kinh doanh"; ws["D3"].font=font_body
    headers=["Ngày","Giờ vào","Giờ ra","Giờ làm","OT (giờ)","Công","Ghi chú"]
    for c,h in enumerate(headers,1): ws.cell(row=5,column=c,value=h)
    style_header_row(ws,5,7)
    days=[("01/05","8:00","17:30",8.5,0,1,""),("02/05","8:15","18:00",8.75,1,1,"OT 1h"),
          ("03/05","8:00","17:00",8.0,0,1,""),("04/05","","",0,0,0,"Nghỉ"),
          ("05/05","8:30","20:00",10.5,2.5,1,"OT 2.5h"),("06/05","8:00","17:30",8.5,0,1,""),
          ("07/05","9:00","17:00",7.0,0,0.875,"Đi trễ 1h")]
    wh=0;ot=0
    for i,(d,come,leave,hours,ot_h,cong,note) in enumerate(days,6):
        wh+=hours;ot+=ot_h
        ws.cell(row=i,column=1,value=d);ws.cell(row=i,column=2,value=come);ws.cell(row=i,column=3,value=leave)
        ws.cell(row=i,column=4,value=hours);ws.cell(row=i,column=5,value=ot_h)
        ws.cell(row=i,column=6,value=cong);ws.cell(row=i,column=7,value=note)
        for c in range(1,8):
            style_body_cell(ws.cell(row=i,column=c),center=True)
            if i%2==0: ws.cell(row=i,column=c).fill=fill_pale
    row_t=6+len(days)
    ws.cell(row=row_t,column=1,value="TỔNG").font=font_body_bold
    for col,val in [(4,wh),(5,ot),(6,sum(d[4] for d in days))]:
        ws.cell(row=row_t,column=col,value=val).font=font_body_bold;ws.cell(row=row_t,column=col).alignment=align_center
    set_col_widths(ws,[12,10,10,12,12,8,18]);ws.sheet_view.showGridLines=False
    wb.save(os.path.join(OUT,"timesheet.xlsx"));print("✅ timesheet.xlsx")

# =====================================================
# 4. Gantt Chart
# =====================================================
def create_gantt():
    wb=openpyxl.Workbook();ws=wb.active;ws.title="Gantt Chart"
    ws.merge_cells("A1:H1");ws["A1"].value="SƠ ĐỒ GANTT - DỰ ÁN ABC"
    ws["A1"].font=font_title;ws["A1"].alignment=align_center;ws.row_dimensions[1].height=36
    headers=["Công việc","Người phụ trách","Bắt đầu","Kết thúc","Số ngày","Tuần 1","Tuần 2","Tuần 3"]
    for c,h in enumerate(headers,1): ws.cell(row=3,column=c,value=h)
    style_header_row(ws,3,8)
    tasks=[("Khảo sát yêu cầu","Nam","01/06","05/06",5,"███","",""),
           ("Phân tích hệ thống","Lan","04/06","10/06",5,"","███",""),
           ("Thiết kế UI/UX","Minh","08/06","14/06",5,"","██","███"),
           ("Phát triển Backend","Hoàng","11/06","20/06",8,"","","████"),
           ("Phát triển Frontend","Minh","11/06","22/06",10,"","","█████"),
           ("Kiểm thử","Lan","21/06","25/06",5,"","","██"),
           ("Triển khai","Nam","25/06","28/06",4,"","","")]
    for i,(task,person,start,end,days,*weeks) in enumerate(tasks,4):
        ws.cell(row=i,column=1,value=task);ws.cell(row=i,column=2,value=person)
        ws.cell(row=i,column=3,value=start);ws.cell(row=i,column=4,value=end);ws.cell(row=i,column=5,value=days)
        for j,w in enumerate(weeks,6): ws.cell(row=i,column=j,value=w if w else "")
        for c in range(1,9): style_body_cell(ws.cell(row=i,column=c),center=(c>=3))
    set_col_widths(ws,[28,16,12,12,10,12,12,12]);ws.sheet_view.showGridLines=False
    wb.save(os.path.join(OUT,"gantt.xlsx"));print("✅ gantt.xlsx")

# =====================================================
# 5. P&L
# =====================================================
def create_pnl():
    wb=openpyxl.Workbook();ws=wb.active;ws.title="Báo cáo P&L"
    ws.merge_cells("A1:D1");ws["A1"].value="BÁO CÁO LỢI NHUẬN (P&L) - QUÝ 2/2025"
    ws["A1"].font=font_title;ws["A1"].alignment=align_center;ws.row_dimensions[1].height=36
    headers=["Chỉ tiêu","Tháng 4","Tháng 5","Tháng 6"]
    for c,h in enumerate(headers,1): ws.cell(row=3,column=c,value=h)
    style_header_row(ws,3,4)
    items=[("DOANH THU",None,None,None),("  Doanh thu bán hàng",120000000,135000000,150000000),
           ("  Doanh thu dịch vụ",45000000,50000000,55000000),("Tổng doanh thu",None,None,None),
           ("CHI PHÍ",None,None,None),("  Nguyên vật liệu",40000000,45000000,48000000),
           ("  Nhân công",30000000,32000000,35000000),("  Vận hành",15000000,15000000,18000000),
           ("  Marketing",10000000,12000000,15000000),("Tổng chi phí",None,None,None),
           ("LỢI NHUẬN GỘP",None,None,None),("  Lợi nhuận",None,None,None),("  Biên lợi nhuận",None,None,None)]
    for i,(item,*vals) in enumerate(items,4):
        ws.cell(row=i,column=1,value=item)
        is_sec=item==item.strip() and item.isupper()
        if is_sec:
            ws.cell(row=i,column=1).font=Font(name="Inter",size=10,bold=True,color=GREEN_DARK)
            for c in range(1,5): ws.cell(row=i,column=c).fill=fill_light
        else: ws.cell(row=i,column=1).font=font_body
        for j,v in enumerate(vals,2):
            cell=ws.cell(row=i,column=j)
            if v is not None: cell.value=v
            cell.alignment=align_center;cell.border=thin_border
            if i%2==0 and not is_sec: cell.fill=fill_pale
    set_col_widths(ws,[30,20,20,20]);ws.sheet_view.showGridLines=False
    wb.save(os.path.join(OUT,"pnl.xlsx"));print("✅ pnl.xlsx")

# =====================================================
# 6. Cash Flow
# =====================================================
def create_cashflow():
    wb=openpyxl.Workbook();ws=wb.active;ws.title="Dòng tiền"
    ws.merge_cells("A1:D1");ws["A1"].value="BẢNG THEO DÕI DÒNG TIỀN - THÁNG 5"
    ws["A1"].font=font_title;ws["A1"].alignment=align_center;ws.row_dimensions[1].height=36
    headers=["Ngày","Mô tả","Tiền vào (VNĐ)","Tiền ra (VNĐ)"]
    for c,h in enumerate(headers,1): ws.cell(row=3,column=c,value=h)
    style_header_row(ws,3,4)
    entries=[("01/05","Số dư đầu kỳ",50000000,""),("02/05","Thu từ khách hàng A",25000000,""),
             ("03/05","Trả tiền văn phòng","",5000000),("05/05","Chi lương nhân viên","",30000000),
             ("07/05","Thu từ khách hàng B",15000000,""),("10/05","Chi mua nguyên vật liệu","",15000000),
             ("12/05","Chi marketing","",8000000),("15/05","Thu từ khách hàng C",12000000,""),
             ("18/05","Thanh toán điện nước","",2000000),("20/05","Chi phí vận chuyển","",3000000),
             ("25/05","Thu khác",5000000,""),("28/05","Chi phí bảo trì","",1500000)]
    for i,(date,desc,inflow,outflow) in enumerate(entries,4):
        ws.cell(row=i,column=1,value=date);ws.cell(row=i,column=2,value=desc)
        ws.cell(row=i,column=3,value=inflow if inflow else "");ws.cell(row=i,column=4,value=outflow if outflow else "")
        for c in range(1,5):
            cell=ws.cell(row=i,column=c);cell.font=font_body;cell.alignment=align_center if c!=2 else align_left
            cell.border=thin_border
            if i%2==0: cell.fill=fill_pale
    br=4+len(entries)
    ws.cell(row=br,column=1,value="TỔNG").font=font_body_bold
    for col in[3,4]:
        cl=get_column_letter(col);ws.cell(row=br,column=col,value=f"=SUM({cl}4:{cl}{br-1})")
        ws.cell(row=br,column=col).font=font_body_bold;ws.cell(row=br,column=col).alignment=align_center
    ws.cell(row=br+1,column=1,value="SỐ DƯ CUỐI KỲ").font=Font(name="Inter",size=11,bold=True,color=GREEN_MID)
    ws.cell(row=br+1,column=3,value=f"={get_column_letter(3)}{br}-{get_column_letter(4)}{br}+50000000")
    ws.cell(row=br+1,column=3).font=Font(name="Inter",size=11,bold=True,color=GREEN_MID);ws.cell(row=br+1,column=3).alignment=align_center
    set_col_widths(ws,[12,32,20,20]);ws.sheet_view.showGridLines=False
    wb.save(os.path.join(OUT,"cashflow.xlsx"));print("✅ cashflow.xlsx")

# =====================================================
# 7. Payroll
# =====================================================
def create_payroll():
    wb=openpyxl.Workbook();ws=wb.active;ws.title="Bảng lương"
    ws.merge_cells("A1:G1");ws["A1"].value="BẢNG TÍNH LƯƠNG THÁNG 5/2025"
    ws["A1"].font=font_title;ws["A1"].alignment=align_center;ws.row_dimensions[1].height=36
    headers=["STT","Họ tên","Lương cơ bản","Phụ cấp","BHXH (8%)","Thuế TNCN","Thực lĩnh"]
    for c,h in enumerate(headers,1): ws.cell(row=3,column=c,value=h)
    style_header_row(ws,3,7)
    emps=[(1,"Nguyễn Văn A",15000000,2000000),(2,"Trần Thị B",12000000,1500000),
          (3,"Lê Văn C",18000000,3000000),(4,"Phạm Thị D",10000000,1000000),(5,"Hoàng Văn E",25000000,5000000)]
    for i,(stt,name,base,allowance) in enumerate(emps,4):
        bhxh=int(base*0.08);thu_nhap=base+allowance-bhxh;thue_tncn=int(thu_nhap*0.05) if thu_nhap>11000000 else 0;thuc_linh=thu_nhap-thue_tncn
        ws.cell(row=i,column=1,value=stt);ws.cell(row=i,column=2,value=name);ws.cell(row=i,column=3,value=base)
        ws.cell(row=i,column=4,value=allowance);ws.cell(row=i,column=5,value=bhxh);ws.cell(row=i,column=6,value=thue_tncn);ws.cell(row=i,column=7,value=thuc_linh)
        for c in range(1,8): style_body_cell(ws.cell(row=i,column=c),center=(c!=2))
    set_col_widths(ws,[6,22,16,14,14,14,16]);ws.sheet_view.showGridLines=False
    wb.save(os.path.join(OUT,"payroll.xlsx"));print("✅ payroll.xlsx")

# =====================================================
# 8. KPI Scorecard
# =====================================================
def create_kpi():
    wb=openpyxl.Workbook();ws=wb.active;ws.title="KPI"
    ws.merge_cells("A1:F1");ws["A1"].value="BẢNG ĐÁNH GIÁ KPI - QUÝ 2/2025"
    ws["A1"].font=font_title;ws["A1"].alignment=align_center;ws.row_dimensions[1].height=36
    headers=["Nhân viên","Chỉ tiêu","Mục tiêu","Kết quả","% Đạt","Xếp loại"]
    for c,h in enumerate(headers,1): ws.cell(row=3,column=c,value=h)
    style_header_row(ws,3,6)
    data=[("Nguyễn Văn A","Doanh số",500000000,520000000,104,"Xuất sắc"),("Nguyễn Văn A","Khách hàng mới",10,12,120,"Vượt trội"),
          ("Trần Thị B","Doanh số",400000000,380000000,95,"Tốt"),("Trần Thị B","Hoàn thành dự án",5,5,100,"Xuất sắc"),
          ("Lê Văn C","Doanh số",600000000,450000000,75,"Cần cải thiện"),("Lê Văn C","Hài lòng khách hàng",90,88,97.8,"Tốt")]
    for i,(name,metric,target,result,pct,rank) in enumerate(data,4):
        ws.cell(row=i,column=1,value=name);ws.cell(row=i,column=2,value=metric);ws.cell(row=i,column=3,value=target)
        ws.cell(row=i,column=4,value=result);ws.cell(row=i,column=5,value=f"{pct}%");ws.cell(row=i,column=6,value=rank)
        for c in range(1,7):
            style_body_cell(ws.cell(row=i,column=c),center=(c>=3))
            if i%2==0: ws.cell(row=i,column=c).fill=fill_pale
    set_col_widths(ws,[22,20,18,18,12,18]);ws.sheet_view.showGridLines=False
    wb.save(os.path.join(OUT,"kpi.xlsx"));print("✅ kpi.xlsx")

# =====================================================
# 9. Task Tracker
# =====================================================
def create_task_tracker():
    wb=openpyxl.Workbook();ws=wb.active;ws.title="Task"
    ws.merge_cells("A1:G1");ws["A1"].value="THEO DÕI CÔNG VIỆC"
    ws["A1"].font=font_title;ws["A1"].alignment=align_center;ws.row_dimensions[1].height=36
    headers=["STT","Công việc","Người phụ trách","Priority","Hạn chót","Trạng thái","% Hoàn thành"]
    for c,h in enumerate(headers,1): ws.cell(row=3,column=c,value=h)
    style_header_row(ws,3,7)
    tasks=[(1,"Báo cáo tuần","Nam","Cao","21/05","Hoàn thành",100),(2,"Thiết kế landing page","Minh","Cao","25/05","Đang làm",60),
           (3,"Họp khách hàng","Lan","Trung bình","22/05","Chờ",0),(4,"Cập nhật hợp đồng","Hoàng","Trung bình","23/05","Đang làm",30),
           (5,"Báo giá thiết bị","Nam","Thấp","28/05","Chưa bắt đầu",0),(6,"Viết blog SEO","Minh","Trung bình","24/05","Đang làm",75),
           (7,"Phân tích đối thủ","Lan","Thấp","30/05","Chưa bắt đầu",0)]
    for i,(stt,task,person,priority,deadline,status,pct) in enumerate(tasks,4):
        ws.cell(row=i,column=1,value=stt);ws.cell(row=i,column=2,value=task);ws.cell(row=i,column=3,value=person)
        ws.cell(row=i,column=4,value=priority);ws.cell(row=i,column=5,value=deadline)
        ws.cell(row=i,column=6,value=status);ws.cell(row=i,column=7,value=pct)
        for c in range(1,8): style_body_cell(ws.cell(row=i,column=c),center=(c!=2))
    set_col_widths(ws,[6,30,16,14,12,16,14]);ws.sheet_view.showGridLines=False
    wb.save(os.path.join(OUT,"task-tracker.xlsx"));print("✅ task-tracker.xlsx")

# =====================================================
# 10-16: More templates
# =====================================================
def create_quotation():
    wb=openpyxl.Workbook();ws=wb.active;ws.title="Báo giá"
    ws.merge_cells("A1:F1");ws["A1"].value="BÁO GIÁ";ws["A1"].font=Font(name="Inter",size=18,bold=True,color=GREEN_DARK);ws["A1"].alignment=align_center;ws.row_dimensions[1].height=44
    ws["A3"].value="Số báo giá: BG-2025-001";ws["A3"].font=font_body_bold;ws["E3"].value="Ngày: 21/05/2025";ws["E3"].font=font_body_bold;ws["E3"].alignment=align_right
    ws["A4"].value="Khách hàng: Công ty TNHH XYZ";ws["A4"].font=font_body
    headers=["STT","Sản phẩm/Dịch vụ","Đơn giá (VNĐ)","Số lượng","Thành tiền","Ghi chú"]
    for c,h in enumerate(headers,1): ws.cell(row=6,column=c,value=h)
    style_header_row(ws,6,6)
    items=[(1,"Máy tính Dell OptiPlex",15000000,5),(2,"Màn hình Dell 27\"",5000000,5),(3,"Bàn làm việc",2000000,5),(4,"Ghế văn phòng",3000000,5)]
    for i,(stt,name,price,qty) in enumerate(items,7):
        ws.cell(row=i,column=1,value=stt);ws.cell(row=i,column=2,value=name);ws.cell(row=i,column=3,value=price)
        ws.cell(row=i,column=4,value=qty);ws.cell(row=i,column=5,value=price*qty);ws.cell(row=i,column=6,value="")
        for c in range(1,7): style_body_cell(ws.cell(row=i,column=c),center=(c!=2))
    set_col_widths(ws,[6,30,18,10,18,20]);ws.sheet_view.showGridLines=False
    wb.save(os.path.join(OUT,"quotation.xlsx"));print("✅ quotation.xlsx")

def create_calendar():
    wb=openpyxl.Workbook();ws=wb.active;ws.title="Lịch"
    ws.merge_cells("A1:G1");ws["A1"].value="LỊCH THÁNG 5/2025";ws["A1"].font=font_title;ws["A1"].alignment=align_center;ws.row_dimensions[1].height=36
    headers=["T2","T3","T4","T5","T6","T7","CN"]
    for c,h in enumerate(headers,1): ws.cell(row=3,column=c,value=h)
    style_header_row(ws,3,7)
    for r in range(6):
        for c in range(7):
            cell=ws.cell(row=4+r,column=c+1)
            style_body_cell(cell,center=True)
    days=[(4,1,"1"),(4,2,"2"),(5,1,"5"),(5,2,"6"),(6,3,"12"),(7,5,"19"),(8,7,"25")]
    for row,col,val in days:
        ws.cell(row=row,column=col,value=val).font=font_body_bold
    set_col_widths(ws,[10]*7);ws.sheet_view.showGridLines=False
    wb.save(os.path.join(OUT,"calendar.xlsx"));print("✅ calendar.xlsx")

def create_sales_tracker():
    wb=openpyxl.W
# ===== Run all =====
if __name__ == "__main__":
    create_budget()
    create_invoice()
    create_timesheet()
    create_gantt()
    create_pnl()
    create_cashflow()
    create_payroll()
    create_kpi()
    create_task_tracker()
    create_quotation()
    create_calendar()
    create_sales_tracker()
    print("✅ All templates generated!")
