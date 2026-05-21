#!/usr/bin/env python3
import os, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = os.path.expanduser("~/Desktop/Openclaw/17_Office_Templates/templates/excel")
os.makedirs(OUT, exist_ok=True)

GREEN_DARK="14532d";GREEN_MID="16a34a";GREEN_LIGHT="dcfce7";GREEN_PALE="f0fdf4";WHITE="ffffff"
font_title=Font(name="Inter",size=14,bold=True,color=GREEN_DARK)
font_header=Font(name="Inter",size=11,bold=True,color=WHITE)
font_body=Font(name="Inter",size=10,color=GREEN_DARK)
font_body_bold=Font(name="Inter",size=10,bold=True,color=GREEN_DARK)
fill_header=PatternFill(start_color=GREEN_MID,end_color=GREEN_MID,fill_type="solid")
fill_pale=PatternFill(start_color=GREEN_PALE,end_color=GREEN_PALE,fill_type="solid")
fill_light=PatternFill(start_color=GREEN_LIGHT,end_color=GREEN_LIGHT,fill_type="solid")
align_center=Alignment(horizontal="center",vertical="center",wrap_text=True)
align_left=Alignment(horizontal="left",vertical="center",wrap_text=True)
thin_border=Border(left=Side(style="thin",color="bbf7d0"),right=Side(style="thin",color="bbf7d0"),top=Side(style="thin",color="bbf7d0"),bottom=Side(style="thin",color="bbf7d0"))
header_border=Border(left=Side(style="thin",color=GREEN_MID),right=Side(style="thin",color=GREEN_MID),top=Side(style="thin",color=GREEN_MID),bottom=Side(style="medium",color=GREEN_MID))

def shr(ws,row,cols):
    for c in range(1,cols+1):
        cell=ws.cell(row=row,column=c);cell.font=font_header;cell.fill=fill_header;cell.alignment=align_center;cell.border=header_border
def sbc(cell,cen=False):
    cell.font=font_body;cell.fill=PatternFill(start_color=WHITE,end_color=WHITE,fill_type="solid");cell.alignment=align_center if cen else align_left;cell.border=thin_border
def scw(ws,widths):
    for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=w

# EXPENSE TRACKER
def create_expense():
    wb=openpyxl.Workbook();ws=wb.active;ws.title="Chi phí"
    ws.merge_cells("A1:E1");ws["A1"].value="THEO DÕI CHI PHÍ";ws["A1"].font=font_title;ws["A1"].alignment=align_center;ws.row_dimensions[1].height=36
    h=["Ngày","Danh mục","Mô tả","Số tiền (VNĐ)","Ghi chú"]
    for c,h in enumerate(h,1): ws.cell(row=3,column=c,value=h)
    shr(ws,3,5)
    items=[("02/05","Ăn uống","Ăn trưa cùng đồng nghiệp",150000,""),("03/05","Di chuyển","Đổ xăng",500000,""),
           ("05/05","Giải trí","Xem phim",200000,""),("06/05","Mua sắm","Áo sơ mi",350000,""),
           ("08/05","Hóa đơn","Tiền điện T5",1200000,""),("10/05","Sức khỏe","Thuốc bổ",250000,""),
           ("12/05","Ăn uống","Đi ăn tối gia đình",600000,""),("15/05","Giáo dục","Khóa học online",800000,"")]
    for i,(date,cat,desc,amt,note) in enumerate(items,4):
        ws.cell(row=i,column=1,value=date);ws.cell(row=i,column=2,value=cat);ws.cell(row=i,column=3,value=desc)
        ws.cell(row=i,column=4,value=amt);ws.cell(row=i,column=5,value=note)
        for c in range(1,6): sbc(ws.cell(row=i,column=c),cen=(c!=3))
    rt=4+len(items);ws.cell(row=rt,column=1,value="TỔNG").font=font_body_bold
    ws.cell(row=rt,column=4,value=f"=SUM(D4:D{rt-1})").font=font_body_bold;ws.cell(row=rt,column=4).alignment=align_center
    scw(ws,[12,14,30,18,15]);ws.sheet_view.showGridLines=False
    wb.save(os.path.join(OUT,"expense.xlsx"));print("✅ expense.xlsx")

# CRM
def create_crm():
    wb=openpyxl.Workbook();ws=wb.active;ws.title="CRM"
    ws.merge_cells("A1:G1");ws["A1"].value="QUẢN LÝ KHÁCH HÀNG";ws["A1"].font=font_title;ws["A1"].alignment=align_center;ws.row_dimensions[1].height=36
    h=["STT","Tên khách hàng","Công ty","SĐT","Email","Trạng thái","Ghi chú"]
    for c,h in enumerate(h,1): ws.cell(row=3,column=c,value=h)
    shr(ws,3,7)
    items=[(1,"Nguyễn Văn A","VMO Holding","0901xxx","a@gmail.com","Tiềm năng","Đã gặp T4"),
           (2,"Trần Thị B","FPT Software","0902xxx","b@gmail.com","Đàm phán","Báo giá xong"),
           (3,"Lê Văn C","VNG Corp","0903xxx","c@gmail.com","Khách quen","Đã ký HĐ"),
           (4,"Phạm Thị D","MISA","0904xxx","d@gmail.com","Mới","Cần gọi lại"),
           (5,"Hoàng Văn E","Viettel","0905xxx","e@gmail.com","Tiềm năng","Giới thiệu")]
    for i,(stt,name,co,phone,email,status,note) in enumerate(items,4):
        ws.cell(row=i,column=1,value=stt);ws.cell(row=i,column=2,value=name);ws.cell(row=i,column=3,value=co)
        ws.cell(row=i,column=4,value=phone);ws.cell(row=i,column=5,value=email);ws.cell(row=i,column=6,value=status);ws.cell(row=i,column=7,value=note)
        for c in range(1,8): sbc(ws.cell(row=i,column=c),cen=(c!=2 and c!=3))
    scw(ws,[6,18,22,14,22,16,20]);ws.sheet_view.showGridLines=False
    wb.save(os.path.join(OUT,"crm.xlsx"));print("✅ crm.xlsx")

# MEETING MINUTES
def create_meeting():
    wb=openpyxl.Workbook();ws=wb.active;ws.title="Biên bản họp"
    ws.merge_cells("A1:F1");ws["A1"].value="BIÊN BẢN CUỘC HỌP";ws["A1"].font=font_title;ws["A1"].alignment=align_center;ws.row_dimensions[1].height=36
    ws["A3"].value="Chủ đề: Họp tiến độ dự án Q2";ws["A3"].font=font_body_bold
    ws["A4"].value="Ngày: 20/05/2025";ws["A4"].font=font_body;ws["D4"].value="Địa điểm: Phòng họp A";ws["D4"].font=font_body
    ws["A5"].value="Người tham dự: Nam, Lan, Minh, Hoàng";ws["A5"].font=font_body
    h=["STT","Nội dung","Người phụ trách","Deadline","Trạng thái","Ghi chú"]
    for c,h in enumerate(h,1): ws.cell(row=7,column=c,value=h)
    shr(ws,7,6)
    items=[(1,"Hoàn thiện thiết kế UI","Minh","25/05","Đang làm","80% xong"),
           (2,"Triển khai API đăng nhập","Hoàng","23/05","Hoàn thành","Đã test OK"),
           (3,"Báo cáo doanh thu Q1","Lan","22/05","Chờ","Cần số liệu từ kế toán"),
           (4,"Khảo sát khách hàng mới","Nam","30/05","Chưa bắt đầu","Gửi email survey")]
    for i,(stt,content,person,deadline,status,note) in enumerate(items,8):
        ws.cell(row=i,column=1,value=stt);ws.cell(row=i,column=2,value=content);ws.cell(row=i,column=3,value=person)
        ws.cell(row=i,column=4,value=deadline);ws.cell(row=i,column=5,value=status);ws.cell(row=i,column=6,value=note)
        for c in range(1,7): sbc(ws.cell(row=i,column=c),cen=(c!=2))
    scw(ws,[6,32,16,12,14,20]);ws.sheet_view.showGridLines=False
    wb.save(os.path.join(OUT,"meeting.xlsx"));print("✅ meeting.xlsx")

# LEAVE FORM
def create_leave():
    wb=openpyxl.Workbook();ws=wb.active;ws.title="Đơn nghỉ phép"
    ws.merge_cells("A1:D1");ws["A1"].value="ĐƠN XIN NGHỈ PHÉP";ws["A1"].font=font_title;ws["A1"].alignment=align_center;ws.row_dimensions[1].height=36
    fields=[("Họ và tên: Nguyễn Văn A",1),("Phòng ban: Kinh doanh",2),("Ngày bắt đầu: 26/05/2025",3),("Ngày kết thúc: 28/05/2025",4)]
    for val,row in fields:
        ws.cell(row=2+row,column=1,value=val).font=font_body
    ws.cell(row=7,column=1,value="Loại phép:").font=font_body_bold
    types=["Nghỉ phép năm ☑","Nghỉ bệnh ☐","Nghỉ việc riêng ☐"]
    for i,t in enumerate(types): ws.cell(row=8+i,column=1,value=t).font=font_body
    ws.cell(row=12,column=1,value="Lý do:").font=font_body_bold
    ws.cell(row=13,column=1,value="Việc gia đình").font=font_body
    ws.cell(row=15,column=1,value="Người phê duyệt:").font=font_body_bold
    ws.cell(row=16,column=1,value="Trưởng phòng").font=font_body
    ws.cell(row=18,column=1,value="Ký tên: ______________").font=Font(name="Inter",size=10,italic=True,color="166534")
    scw(ws,[40,20,20,20]);ws.sheet_view.showGridLines=False
    wb.save(os.path.join(OUT,"leave-form.xlsx"));print("✅ leave-form.xlsx")

# HABIT TRACKER
def create_habit():
    wb=openpyxl.Workbook();ws=wb.active;ws.title="Thói quen"
    ws.merge_cells("A1:H1");ws["A1"].value="HABIT TRACKER - THÁNG 5";ws["A1"].font=font_title;ws["A1"].alignment=align_center;ws.row_dimensions[1].height=36
    h=["Thói quen","T2","T3","T4","T5","T6","T7","CN"]
    for c,h in enumerate(h,1): ws.cell(row=3,column=c,value=h)
    shr(ws,3,8)
    habits=[("Tập thể dục","✅","✅","✅","","✅","✅",""),("Đọc sách 30 phút","✅","✅","","✅","✅","✅","✅"),
            ("Uống 2L nước","✅","✅","✅","✅","✅","✅","✅"),("Thiền 10 phút","✅","","✅","✅","","✅","")]
    for i,(habit,*days) in enumerate(habits,4):
        ws.cell(row=i,column=1,value=habit)
        for j,d in enumerate(days,2): ws.cell(row=i,column=j,value=d)
        for c in range(1,9): sbc(ws.cell(row=i,column=c),cen=True)
    scw(ws,[20,6,6,6,6,6,6,6]);ws.sheet_view.showGridLines=False
    wb.save(os.path.join(OUT,"habit.xlsx"));print("✅ habit.xlsx")

# LOAN CALCULATOR
def create_loan():
    wb=openpyxl.Workbook();ws=wb.active;ws.title="Vay"
    ws.merge_cells("A1:E1");ws["A1"].value="BẢNG TÍNH LÃI VAY";ws["A1"].font=font_title;ws["A1"].alignment=align_center;ws.row_dimensions[1].height=36
    ws["A3"].value="Số tiền vay: 500,000,000 VNĐ";ws["A3"].font=font_body_bold;ws["D3"].value="Lãi suất: 8%/năm";ws["D3"].font=font_body_bold
    ws["A4"].value="Thời hạn: 60 tháng";ws["A4"].font=font_body;ws["D4"].value="Kỳ hạn: Hàng tháng";ws["D4"].font=font_body
    h=["Kỳ","Dư nợ đầu kỳ","Tiền gốc","Tiền lãi","Tổng trả"]
    for c,h in enumerate(h,1): ws.cell(row=6,column=c,value=h)
    shr(ws,6,5)
    for i in range(6):
        row=7+i;ws.cell(row=row,column=1,value=i+1);ws.cell(row=row,column=2,value=500000000-i*8333333)
        ws.cell(row=row,column=3,value=8333333);ws.cell(row=row,column=4,value=round((500000000-i*8333333)*0.08/12))
        ws.cell(row=row,column=5,value=8333333+round((500000000-i*8333333)*0.08/12))
        for c in range(1,6): sbc(ws.cell(row=row,column=c),cen=True)
    scw(ws,[8,18,14,14,14]);ws.sheet_view.showGridLines=False
    wb.save(os.path.join(OUT,"loan.xlsx"));print("✅ loan.xlsx")

# SALES TRACKER
def create_sales():
    wb=openpyxl.Workbook();ws=wb.active;ws.title="Doanh số"
    ws.merge_cells("A1:G1");ws["A1"].value="THEO DÕI DOANH SỐ BÁN HÀNG";ws["A1"].font=font_title;ws["A1"].alignment=align_center;ws.row_dimensions[1].height=36
    h=["Tuần","Mục tiêu","Doanh số","% Đạt","Đơn hàng","Khách mới","Ghi chú"]
    for c,h in enumerate(h,1): ws.cell(row=3,column=c,value=h)
    shr(ws,3,7)
    data=[("Tuần 1",50000000,45000000,90,15,5,"Còn thiếu 5tr"),("Tuần 2",50000000,55000000,110,18,7,"Vượt 10%"),
          ("Tuần 3",50000000,48000000,96,14,4,"Gần đạt"),("Tuần 4",50000000,62000000,124,20,8,"Vượt 24%")]
    for i,(wk,target,actual,pct,orders,new,note) in enumerate(data,4):
        ws.cell(row=i,column=1,value=wk);ws.cell(row=i,column=2,value=target);ws.cell(row=i,column=3,value=actual)
        ws.cell(row=i,column=4,value=f"{pct}%");ws.cell(row=i,column=5,value=orders);ws.cell(row=i,column=6,value=new);ws.cell(row=i,column=7,value=note)
        for c in range(1,8): sbc(ws.cell(row=i,column=c),cen=(c!=1))
    scw(ws,[10,16,16,10,12,12,20]);ws.sheet_view.showGridLines=False
    wb.save(os.path.join(OUT,"sales-tracker.xlsx"));print("✅ sales-tracker.xlsx")

# GRADE TRACKER
def create_grade():
    wb=openpyxl.Workbook();ws=wb.active;ws.title="Điểm"
    ws.merge_cells("A1:F1");ws["A1"].value="BẢNG ĐIỂM HỌC SINH";ws["A1"].font=font_title;ws["A1"].alignment=align_center;ws.row_dimensions[1].height=36
    h=["STT","Họ tên","Toán","Văn","Anh","TB"]
    for c,h in enumerate(h,1): ws.cell(row=3,column=c,value=h)
    shr(ws,3,6)
    items=[(1,"Nguyễn Văn A",8.5,7.0,9.0),(2,"Trần Thị B",6.5,8.0,7.5),(3,"Lê Văn C",9.0,6.5,8.5),
           (4,"Phạm Thị D",7.0,8.5,7.0),(5,"Hoàng Văn E",5.5,6.0,6.5)]
    for i,(stt,name,toan,van,anh) in enumerate(items,4):
        tb=round((toan+van+anh)/3,1)
        ws.cell(row=i,column=1,value=stt);ws.cell(row=i,column=2,value=name);ws.cell(row=i,column=3,value=toan)
        ws.cell(row=i,column=4,value=van);ws.cell(row=i,column=5,value=anh);ws.cell(row=i,column=6,value=tb)
        for c in range(1,7): sbc(ws.cell(row=i,column=c),cen=(c!=2))
    scw(ws,[6,22,8,8,8,8]);ws.sheet_view.showGridLines=False
    wb.save(os.path.join(OUT,"grade.xlsx"));print("✅ grade.xlsx")

# CONTENT CALENDAR
def create_content():
    wb=openpyxl.Workbook();ws=wb.active;ws.title="Content"
    ws.merge_cells("A1:G1");ws["A1"].value="LỊCH CONTENT MARKETING - THÁNG 6";ws["A1"].font=font_title;ws["A1"].alignment=align_center;ws.row_dimensions[1].height=36
    h=["Ngày","Nền tảng","Nội dung","Định dạng","Trạng thái","Người viết","Engagement"]
    for c,h in enumerate(h,1): ws.cell(row=3,column=c,value=h)
    shr(ws,3,7)
    items=[("01/06","Facebook","Giới thiệu sản phẩm mới","Bài viết","Đã lên lịch","Minh",""),
           ("03/06","Instagram","Behind the scenes","Video","Đang làm","Lan",""),
           ("05/06","LinkedIn","Case study khách hàng","Bài viết","Lên ý tưởng","Nam",""),
           ("08/06","TikTok","Review sản phẩm","Video ngắn","Đang quay","Minh",""),
           ("10/06","Facebook","Khuyến mãi tháng 6","Bài viết","Đã lên lịch","Lan",""),
           ("12/06","Instagram","Q&A với founder","Live","Lên kế hoạch","Nam","")]
    for i,(date,plat,content,format,status,author,eng) in enumerate(items,4):
        ws.cell(row=i,column=1,value=date);ws.cell(row=i,column=2,value=plat);ws.cell(row=i,column=3,value=content)
        ws.cell(row=i,column=4,value=format);ws.cell(row=i,column=5,value=status);ws.cell(row=i,column=6,value=author);ws.cell(row=i,column=7,value=eng)
        for c in range(1,8): sbc(ws.cell(row=i,column=c),cen=(c!=3))
    scw(ws,[10,14,30,14,14,12,14]);ws.sheet_view.showGridLines=False
    wb.save(os.path.join(OUT,"content-calendar.xlsx"));print("✅ content-calendar.xlsx")

# INVENTORY
def create_inventory():
    wb=openpyxl.Workbook();ws=wb.active;ws.title="Kho"
    ws.merge_cells("A1:G1");ws["A1"].value="QUẢN LÝ KHO HÀNG";ws["A1"].font=font_title;ws["A1"].alignment=align_center;ws.row_dimensions[1].height=36
    h=["STT","Mã hàng","Tên sản phẩm","Tồn đầu","Nhập","Xuất","Tồn cuối"]
    for c,h in enumerate(h,1): ws.cell(row=3,column=c,value=h)
    shr(ws,3,7)
    items=[(1,"SP001","Áo thun nam",50,100,80),(2,"SP002","Quần jean",30,50,40),(3,"SP003","Váy nữ",20,60,50),
           (4,"SP004","Giày thể thao",15,30,25),(5,"SP005","Túi xách",10,20,15)]
    for i,(stt,code,name,start,inp,outp) in enumerate(items,4):
        end=start+inp-outp
        ws.cell(row=i,column=1,value=stt);ws.cell(row=i,column=2,value=code);ws.cell(row=i,column=3,value=name)
        ws.cell(row=i,column=4,value=start);ws.cell(row=i,column=5,value=inp);ws.cell(row=i,column=6,value=outp);ws.cell(row=i,column=7,value=end)
        for c in range(1,8): sbc(ws.cell(row=i,column=c),cen=(c!=3))
    scw(ws,[6,10,22,10,10,10,10]);ws.sheet_view.showGridLines=False
    wb.save(os.path.join(OUT,"inventory.xlsx"));print("✅ inventory.xlsx")

# SCHEDULE
def create_schedule():
    wb=openpyxl.Workbook();ws=wb.active;ws.title="Lịch tuần"
    ws.merge_cells("A1:G1");ws["A1"].value="LỊCH LÀM VIỆC TUẦN";ws["A1"].font=font_title;ws["A1"].alignment=align_center;ws.row_dimensions[1].height=36
    h=["Giờ","T2","T3","T4","T5","T6","T7"]
    for c,h in enumerate(h,1): ws.cell(row=3,column=c,value=h)
    shr(ws,3,7)
    items=[("8:00-9:00","Họp team","Gọi khách","Họp dự án","Báo cáo","Gọi khách",""),
           ("9:00-10:00","Code","Meeting","Phát triển","Training","Phát triển",""),
           ("10:00-12:00","Làm việc","Làm việc","Làm việc","Làm việc","Làm việc",""),
           ("13:00-15:00","Họp KH","Thiết kế","Viết tài liệu","Demo","Báo cáo tuần",""),
           ("15:00-17:00","Làm việc","Làm việc","Review","Làm việc","Kết thúc tuần","")]
    for i,(time,*slots) in enumerate(items,4):
        ws.cell(row=i,column=1,value=time)
        for j,s in enumerate(slots,2): ws.cell(row=i,column=j,value=s)
        for c in range(1,8): sbc(ws.cell(row=i,column=c))
    scw(ws,[14,14,14,14,14,14,14]);ws.sheet_view.showGridLines=False
    wb.save(os.path.join(OUT,"schedule.xlsx"));print("✅ schedule.xlsx")

# STUDY PLANNER
def create_study():
    wb=openpyxl.Workbook();ws=wb.active;ws.title="Kế hoạch học tập"
    ws.merge_cells("A1:F1");ws["A1"].value="KẾ HOẠCH HỌC TẬP";ws["A1"].font=font_title;ws["A1"].alignment=align_center;ws.row_dimensions[1].height=36
    h=["Môn","Thứ","Giờ","Địa điểm","Mục tiêu","Trạng thái"]
    for c,h in enumerate(h,1): ws.cell(row=3,column=c,value=h)
    shr(ws,3,6)
    items=[("Tiếng Anh","T2,T4,T6","7:00-8:00","Online Zoom","Hoàn thành Unit 5","Đang học"),
           ("Excel","T3,T5","19:00-20:30","Tại nhà","Xong PivotTable","Chưa bắt đầu"),
           ("Marketing","T7","9:00-11:00","Thư viện","Đọc chương 3","Đang học"),
           ("Kỹ năng mềm","CN","14:00-16:00","Online","Khóa communication","Đã xong")]
    for i,(subj,day,time,place,goal,status) in enumerate(items,4):
        ws.cell(row=i,column=1,value=subj);ws.cell(row=i,column=2,value=day);ws.cell(row=i,column=3,value=time)
        ws.cell(row=i,column=4,value=place);ws.cell(row=i,column=5,value=goal);ws.cell(row=i,column=6,value=status)
        for c in range(1,7): sbc(ws.cell(row=i,column=c))
    scw(ws,[16,16,16,16,30,16]);ws.sheet_view.showGridLines=False
    wb.save(os.path.join(OUT,"study-planner.xlsx"));print("✅ study-planner.xlsx")

# RISK
def create_risk():
    wb=openpyxl.Workbook();ws=wb.active;ws.title="Rủi ro"
    ws.merge_cells("A1:F1");ws["A1"].value="MA TRẬN RỦI RO";ws["A1"].font=font_title;ws["A1"].alignment=align_center;ws.row_dimensions[1].height=36
    h=["Rủi ro","Xác suất","Tác động","Mức độ","Giải pháp","Người phụ trách"]
    for c,h in enumerate(h,1): ws.cell(row=3,column=c,value=h)
    shr(ws,3,6)
    items=[("Chậm tiến độ dự án","Cao","Cao","Nguy hiểm","Tăng nguồn lực","Nam"),
           ("Thiếu nhân sự","Trung bình","Cao","Cao","Tuyển thêm","HR"),
           ("Lỗi kỹ thuật","Thấp","Cao","Trung bình","Test kỹ","Hoàng"),
           ("Ngân sách vượt","Trung bình","Trung bình","Trung bình","Kiểm soát chi","Lan"),
           ("Khách hàng hủy","Thấp","Thấp","Thấp","Hợp đồng rõ","Minh")]
    for i,(risk,prob,impact,level,solution,owner) in enumerate(items,4):
        ws.cell(row=i,column=1,value=risk);ws.cell(row=i,column=2,value=prob);ws.cell(row=i,column=3,value=impact)
        ws.cell(row=i,column=4,value=level);ws.cell(row=i,column=5,value=solution);ws.cell(row=i,column=6,value=owner)
        for c in range(1,7): sbc(ws.cell(row=i,column=c))
    scw(ws,[24,14,14,14,24,16]);ws.sheet_view.showGridLines=False
    wb.save(os.path.join(OUT,"risk.xlsx"));print("✅ risk.xlsx")

# VENDOR
def create_vendor():
    wb=openpyxl.Workbook();ws=wb.active;ws.title="Nhà cung cấp"
    ws.merge_cells("A1:G1");ws["A1"].value="THEO DÕI NHÀ CUNG CẤP";ws["A1"].font=font_title;ws["A1"].alignment=align_center;ws.row_dimensions[1].height=36
    h=["STT","Nhà cung cấp","Sản phẩm","SĐT","Email","Đánh giá","Ghi chú"]
    for c,h in enumerate(h,1): ws.cell(row=3,column=c,value=h)
    shr(ws,3,7)
    items=[(1,"Công ty X","Nguyên vật liệu","0901xxx","x@gmail.com","Tốt","Đã hợp tác"),
           (2,"Công ty Y","Văn phòng phẩm","0902xxx","y@gmail.com","Rất tốt","Giá tốt"),
           (3,"Công ty Z","Thiết bị CNTT","0903xxx","z@gmail.com","Trung bình","Cần đàm phán lại"),
           (4,"Công ty W","Logistics","0904xxx","w@gmail.com","Tốt","Giao hàng đúng hẹn")]
    for i,(stt,name,product,phone,email,rating,note) in enumerate(items,4):
        ws.cell(row=i,column=1,value=stt);ws.cell(row=i,column=2,value=name);ws.cell(row=i,column=3,value=product)
        ws.cell(row=i,column=4,value=phone);ws.cell(row=i,column=5,value=email);ws.cell(row=i,column=6,value=rating);ws.cell(row=i,column=7,value=note)
        for c in range(1,8): sbc(ws.cell(row=i,column=c))
    scw(ws,[6,20,20,14,24,14,20]);ws.sheet_view.showGridLines=False
    wb.save(os.path.join(OUT,"vendor.xlsx"));print("✅ vendor.xlsx")

# FITNESS
def create_fitness():
    wb=openpyxl.Workbook();ws=wb.active;ws.title="Fitness"
    ws.merge_cells("A1:F1");ws["A1"].value="LỊCH TẬP LUYỆN - THÁNG 5";ws["A1"].font=font_title;ws["A1"].alignment=align_center;ws.row_dimensions[1].height=36
    h=["Ngày","Bài tập","Thời gian","Calo","Cảm giác","Ghi chú"]
    for c,h in enumerate(h,1): ws.cell(row=3,column=c,value=h)
    shr(ws,3,6)
    items=[("02/05","Chạy bộ","30 phút",300,"Tốt","Sáng sớm"),
           ("04/05","Yoga","45 phút",200,"Rất tốt","Thư giãn"),
           ("06/05","Tập tạ","40 phút",350,"Mệt","Tăng độ nặng"),
           ("08/05","Bơi lội","30 phút",250,"Tốt","Buổi chiều"),
           ("10/05","HIIT","20 phút",400,"Rất mệt","Cường độ cao"),
           ("12/05","Đi bộ","60 phút",200,"Nhẹ nhàng","Công viên")]
    for i,(date,ex,time,cal,feel,note) in enumerate(items,4):
        ws.cell(row=i,column=1,value=date);ws.cell(row=i,column=2,value=ex);ws.cell(row=i,column=3,value=time)
    shr(ws,3,6)
    items=[("Nhân sự",50000000,45000000,5000000,90,"Còn 1 tháng"),("Thiết bị",20000000,22000000,-2000000,110,"Phát sinh"),
           ("Văn phòng",10000000,8000000,2000000,80,"Tiết kiệm"),("Marketing",15000000,12000000,3000000,80,"Chưa hết"),
           ("Khác",5000000,3000000,2000000,60,"Dự phòng")]
    for i,(item,budget,actual,diff,pct,note) in enumerate(items,4):
        ws.cell(row=i,column=1,value=item);ws.cell(row=i,column=2,value=budget);ws.cell(row=i,column=3,value=actual)
        ws.cell(row=i,column=4,value=diff);ws.cell(row=i,column=5,value=f"{pct}%");ws.cell(row=i,column=6,value=note)
        for c in range(1,7): sbc(ws.cell(row=i,column=c))
    scw(ws,[20,18,18,18,14,20]);ws.sheet_view.showGridLines=False
    wb.save(os.path.join(OUT,"project-budget.xlsx"));print("✅ project-budget.xlsx")

# TODO
def create_todo():
    wb=openpyxl.Workbook();ws=wb.active;ws.title="Todo"
    ws.merge_cells("A1:D1");ws["A1"].value="DANH SÁCH VIỆC CẦN LÀM";ws["A1"].font=font_title;ws["A1"].alignment=align_center;ws.row_dimensions[1].height=36
    h=["☑","Công việc","Priority","Hạn"]
    for c,h in enumerate(h,1): ws.cell(row=3,column=c,value=h)
    shr(ws,3,4)
    items=[("☐","Gửi báo cáo tuần","Cao","22/05"),("☐","Mua quà sinh nhật","Trung bình","25/05"),
           ("☑","Đặt vé máy bay","Cao","20/05"),("☐","Họp phụ huynh","Cao","28/05"),
           ("☐","Đọc sách 'Atomic Habits'","Thấp","30/06"),("☐","Tập thể dục","Trung bình","Hàng ngày"),
           ("☑","Thanh toán hóa đơn","Cao","15/05")]
    for i,(check,task,prio,deadline) in enumerate(items,4):
        ws.cell(row=i,column=1,value=check);ws.cell(row=i,column=2,value=task)
        ws.cell(row=i,column=3,value=prio);ws.cell(row=i,column=4,value=deadline)
        for c in range(1,5): sbc(ws.cell(row=i,column=c))
    scw(ws,[6,36,14,12]);ws.sheet_view.showGridLines=False
    wb.save(os.path.join(OUT,"todo.xlsx"));print("✅ todo.xlsx")

# ===== RUN ALL =====
if __name__=="__main__":
    from generate_more import *
    create_expense()
    create_crm()
    create_meeting()
    create_leave()
    create_habit()
    create_loan()
    create_sales()
    create_grade()
    create_content()
    create_inventory()
    create_schedule()
    create_study()
    create_risk()
    create_vendor()
    create_fitness()
    create_project_budget()
    create_todo()
    print("✅ All additional templates generated!")
